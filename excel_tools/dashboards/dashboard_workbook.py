"""
dashboard_workbook.py — ADVANCED
==================================
Build a multi-sheet Excel dashboard workbook from a sales CSV:

  Sheet 1 — "Raw Data"    : all records with styled header and auto-sized columns
  Sheet 2 — "Summary"     : aggregated totals by Region (revenue, units)
                            + a native Excel bar chart

Example commands:
    python excel_tools/dashboards/dashboard_workbook.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/dashboard.xlsx

    python excel_tools/dashboards/dashboard_workbook.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/dashboard.xlsx \\
        --group-by region \\
        --force
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Self-contained helpers
# ---------------------------------------------------------------------------

def _coerce(value: str) -> Any:
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def _write_header(ws: Worksheet, headers: list[str], bg: str = "1F4E79") -> None:
    fill = PatternFill(fill_type="solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF")
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws["A2"]


def _autosize(ws: Worksheet, min_w: int = 10, max_w: int = 50) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def _apply_color_scale(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}", rule)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_sales_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        headers = next(reader)
        rows = [[_coerce(v) for v in row] for row in reader]
    return headers, rows


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_raw_data_sheet(ws: Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    """Write all raw data to the sheet with a styled header."""
    _write_header(ws, headers, bg="1F4E79")

    alt_fill = PatternFill(fill_type="solid", fgColor="EDF2FB")
    for r_offset, row in enumerate(rows, start=1):
        r_idx = r_offset + 1
        row_fill = alt_fill if r_offset % 2 == 0 else None
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if row_fill:
                cell.fill = row_fill

    # Conditional format on revenue column if present
    if "revenue" in headers:
        rev_col = headers.index("revenue") + 1
        _apply_color_scale(ws, get_column_letter(rev_col), start_row=2, end_row=len(rows) + 1)

    _autosize(ws)


def build_summary_sheet(
    ws: Worksheet,
    headers: list[str],
    rows: list[list[Any]],
    group_by: str,
) -> None:
    """Aggregate data by *group_by* column and write summary + chart."""
    if group_by not in headers:
        raise ValueError(f"Group-by column '{group_by}' not found. Available: {headers}")

    grp_idx = headers.index(group_by)

    # Identify numeric columns
    numeric_cols: list[tuple[int, str]] = []
    for c_idx, h in enumerate(headers):
        if c_idx == grp_idx:
            continue
        sample_vals = [row[c_idx] for row in rows[:5] if c_idx < len(row)]
        if sample_vals and all(isinstance(v, (int, float)) for v in sample_vals):
            numeric_cols.append((c_idx, h))

    # Aggregate: sum numeric columns per group value
    group_sums: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in rows:
        group_val = str(row[grp_idx]) if grp_idx < len(row) else "Unknown"
        for c_idx, col_name in numeric_cols:
            val = row[c_idx] if c_idx < len(row) else 0
            if isinstance(val, (int, float)):
                group_sums[group_val][col_name] += val

    groups = sorted(group_sums.keys())

    # Write summary header
    summary_headers = [group_by.title()] + [name.replace("_", " ").title() for _, name in numeric_cols]
    _write_header(ws, summary_headers, bg="2E75B6")

    # Write aggregated rows
    for r_offset, group_val in enumerate(groups, start=1):
        r_idx = r_offset + 1
        ws.cell(row=r_idx, column=1, value=group_val)
        for c_offset, (_, col_name) in enumerate(numeric_cols, start=2):
            total = group_sums[group_val][col_name]
            # Round floats to 2 decimal places for cleanliness
            ws.cell(row=r_idx, column=c_offset, value=round(total, 2))

    data_end_row = len(groups) + 1

    # Colour-scale on the last numeric summary column (revenue if present)
    if len(numeric_cols) >= 1:
        rev_summary_col = len(numeric_cols) + 1  # last numeric column (1-indexed)
        _apply_color_scale(ws, get_column_letter(rev_summary_col), 2, data_end_row)

    _autosize(ws)

    # Add bar chart anchored after the data
    if numeric_cols:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Total Revenue by {group_by.title()}"
        chart.style = 10
        chart.width = 22
        chart.height = 14
        chart.y_axis.title = "Revenue"
        chart.x_axis.title = group_by.title()

        # Use the last numeric column (revenue) for the chart series
        rev_col_idx = len(numeric_cols) + 1
        data_ref = Reference(ws, min_col=rev_col_idx, min_row=1, max_row=data_end_row)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=data_end_row)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Place chart to the right of the data
        anchor_col = get_column_letter(len(summary_headers) + 2)
        ws.add_chart(chart, f"{anchor_col}2")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a multi-sheet dashboard workbook from a sales CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python dashboard_workbook.py --input sample_sales.csv --output dashboard.xlsx\n"
            "  python dashboard_workbook.py --input sample_sales.csv "
            "--output dashboard.xlsx --group-by product\n"
        ),
    )
    parser.add_argument("--input", "-i", required=True, type=Path, help="Input CSV file.")
    parser.add_argument("--output", "-o", required=True, type=Path, help="Output .xlsx file.")
    parser.add_argument(
        "--group-by",
        default="region",
        metavar="COLUMN",
        help="Column to group/aggregate by in the Summary sheet (default: region).",
    )
    parser.add_argument("--force", action="store_true", help="Overwrite output if it exists.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if args.output.exists() and not args.force:
        print(f"Error: {args.output} already exists. Use --force to overwrite.", file=sys.stderr)
        return 1

    headers, rows = load_sales_csv(args.input)
    print(f"Loaded {len(rows)} rows × {len(headers)} columns from {args.input}")

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Raw Data
    raw_ws = wb.create_sheet(title="Raw Data")
    build_raw_data_sheet(raw_ws, headers, rows)
    print("Built 'Raw Data' sheet")

    # Sheet 2: Summary
    summary_ws = wb.create_sheet(title="Summary")
    try:
        build_summary_sheet(summary_ws, headers, rows, group_by=args.group_by)
    except ValueError as exc:
        print(f"Error building Summary sheet: {exc}", file=sys.stderr)
        return 1
    print(f"Built 'Summary' sheet (grouped by '{args.group_by}')")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Dashboard workbook written → {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
