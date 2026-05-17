"""
executive_workbook.py — SHOWCASE
==================================
Build a polished, presentation-quality executive workbook from a sales CSV.

Sheets produced:
  1. "Cover"   — branded cover with company name, report title, date, and top-level KPI cells
  2. "Summary" — per-region aggregated KPIs with conditional formatting and a bar chart
  3. "Detail"  — full dataset with a styled header, alternating rows, and column auto-sizing

Design choices:
  - Consistent navy/gold colour palette throughout
  - Merged KPI banner cells on the Cover sheet
  - Named font sizes to communicate information hierarchy
  - Conditional colour-scale on revenue column

Example commands:
    python excel_tools/dashboards/executive_workbook.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/executive_report.xlsx

    python excel_tools/dashboards/executive_workbook.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/executive_report.xlsx \\
        --company "Acme Corp" \\
        --report-title "Q1-Q3 2024 Sales Performance" \\
        --force
"""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Palette constants
# ---------------------------------------------------------------------------
NAVY = "1F4E79"
LIGHT_NAVY = "2F5496"
GOLD = "C9A829"
LIGHT_GOLD = "FFF2CC"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
ALT_ROW = "EDF2FB"
DARK_TEXT = "1F2D3D"
GREEN = "63BE7B"
YELLOW = "FFEB84"
RED = "F8696B"


# ---------------------------------------------------------------------------
# Self-contained style helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_cell(
    ws: Worksheet, row: int, col: int, value: Any,
    bg: str = NAVY, fg: str = WHITE, size: int = 11, bold: bool = True,
    align: str = "center",
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _data_cell(
    ws: Worksheet, row: int, col: int, value: Any,
    bold: bool = False, align: str = "left", fill: PatternFill | None = None,
    number_format: str = "General",
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=DARK_TEXT, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _thin_border()
    if fill:
        cell.fill = fill
    cell.number_format = number_format


def _autosize(ws: Worksheet, min_w: int = 10, max_w: int = 50) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


# ---------------------------------------------------------------------------
# Data loading & aggregation
# ---------------------------------------------------------------------------

def _coerce(v: str) -> Any:
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        return v


def load_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        headers = next(reader)
        rows = [[_coerce(v) for v in row] for row in reader]
    return headers, rows


def aggregate_by_region(
    headers: list[str], rows: list[list[Any]]
) -> tuple[list[str], list[list[Any]]]:
    """Return (agg_headers, agg_rows) aggregated by region column."""
    if "region" not in headers:
        raise ValueError("CSV must contain a 'region' column for aggregation.")
    reg_idx = headers.index("region")
    rev_idx = headers.index("revenue") if "revenue" in headers else None
    unit_idx = headers.index("units") if "units" in headers else None

    totals: dict[str, dict[str, float]] = defaultdict(lambda: {"revenue": 0.0, "units": 0.0, "txn": 0})
    for row in rows:
        region = str(row[reg_idx])
        if rev_idx is not None and rev_idx < len(row):
            totals[region]["revenue"] += float(row[rev_idx]) if isinstance(row[rev_idx], (int, float)) else 0
        if unit_idx is not None and unit_idx < len(row):
            totals[region]["units"] += float(row[unit_idx]) if isinstance(row[unit_idx], (int, float)) else 0
        totals[region]["txn"] += 1

    agg_headers = ["Region", "Transactions", "Total Units", "Total Revenue", "Avg Revenue / Txn"]
    agg_rows: list[list[Any]] = []
    for region in sorted(totals.keys()):
        rev = round(totals[region]["revenue"], 2)
        txn = int(totals[region]["txn"])
        units = int(totals[region]["units"])
        avg = round(rev / txn, 2) if txn else 0.0
        agg_rows.append([region, txn, units, rev, avg])

    return agg_headers, agg_rows


# ---------------------------------------------------------------------------
# Sheet 1: Cover
# ---------------------------------------------------------------------------

def build_cover_sheet(
    ws: Worksheet,
    company: str,
    report_title: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    """Build the branded cover sheet with KPI summary blocks."""
    ws.sheet_view.showGridLines = False

    # Set column widths
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 18
    for r in range(1, 30):
        ws.row_dimensions[r].height = 22

    # --- Company banner (rows 1-3) ---
    ws.merge_cells("B1:G3")
    banner = ws["B1"]
    banner.value = company.upper()
    banner.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    banner.font = Font(bold=True, color=GOLD, size=28)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 40

    # --- Report title (rows 5-6) ---
    ws.merge_cells("B5:G6")
    title_cell = ws["B5"]
    title_cell.value = report_title
    title_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_NAVY)
    title_cell.font = Font(bold=True, color=WHITE, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[5].height = 35
    ws.row_dimensions[6].height = 35

    # --- Date (row 8) ---
    ws.merge_cells("B8:G8")
    date_cell = ws["B8"]
    date_cell.value = f"Prepared: {date.today().strftime('%B %d, %Y')}"
    date_cell.font = Font(color=DARK_TEXT, size=11, italic=True)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- KPI calculations ---
    rev_idx = headers.index("revenue") if "revenue" in headers else None
    unit_idx = headers.index("units") if "units" in headers else None
    reg_idx = headers.index("region") if "region" in headers else None

    total_rev = sum(
        float(r[rev_idx]) for r in rows
        if rev_idx is not None and isinstance(r[rev_idx], (int, float))
    )
    total_units = sum(
        int(r[unit_idx]) for r in rows
        if unit_idx is not None and isinstance(r[unit_idx], (int, float))
    )
    total_txns = len(rows)
    num_regions = len(set(str(r[reg_idx]) for r in rows)) if reg_idx is not None else 0

    kpis = [
        ("Total Revenue",    f"${total_rev:,.2f}",    NAVY),
        ("Total Units Sold", f"{total_units:,}",       LIGHT_NAVY),
        ("Transactions",     f"{total_txns:,}",        "2E75B6"),
        ("Regions",          f"{num_regions}",         "4472C4"),
    ]

    # --- KPI blocks (rows 10-14, each pair of columns) ---
    kpi_start_col = 2  # column B
    for i, (label, value, bg_color) in enumerate(kpis):
        actual_col = kpi_start_col + i

        ws.row_dimensions[10].height = 18
        ws.row_dimensions[11].height = 30
        ws.row_dimensions[12].height = 40
        ws.row_dimensions[13].height = 18

        # Label row
        label_cell = ws.cell(row=11, column=actual_col, value=label)
        label_cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
        label_cell.font = Font(bold=True, color=WHITE, size=10)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.border = _thin_border()

        # Value row
        value_cell = ws.cell(row=12, column=actual_col, value=value)
        value_cell.fill = PatternFill(fill_type="solid", fgColor=LIGHT_GOLD)
        value_cell.font = Font(bold=True, color=NAVY, size=14)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        value_cell.border = _thin_border()

    # --- Footer note (row 16) ---
    ws.merge_cells("B16:G16")
    footer = ws["B16"]
    footer.value = "Confidential — For internal use only"
    footer.font = Font(color="808080", size=9, italic=True)
    footer.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Sheet 2: Summary
# ---------------------------------------------------------------------------

def build_summary_sheet(
    ws: Worksheet,
    agg_headers: list[str],
    agg_rows: list[list[Any]],
) -> None:
    """Build the per-region summary with conditional formatting and chart."""
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(agg_headers))}1")
    title = ws["A1"]
    title.value = "Regional Performance Summary"
    title.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    title.font = Font(bold=True, color=GOLD, size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row (row 2)
    for c_idx, h in enumerate(agg_headers, start=1):
        _header_cell(ws, 2, c_idx, h, bg=LIGHT_NAVY, size=11)
    ws.freeze_panes = ws["A3"]

    # Data rows
    alt_fill = PatternFill(fill_type="solid", fgColor=ALT_ROW)
    for r_offset, row in enumerate(agg_rows, start=1):
        r_idx = r_offset + 2
        row_fill = alt_fill if r_offset % 2 == 0 else None
        for c_idx, value in enumerate(row, start=1):
            _data_cell(
                ws, r_idx, c_idx, value,
                bold=(c_idx == 1),
                align="right" if isinstance(value, (int, float)) else "left",
                fill=row_fill,
                number_format="#,##0.00" if isinstance(value, float) else "General",
            )

    data_end = len(agg_rows) + 2

    # Totals row
    totals_fill = PatternFill(fill_type="solid", fgColor=LIGHT_GOLD)
    _data_cell(ws, data_end + 1, 1, "TOTAL", bold=True, fill=totals_fill)
    for c_idx in range(2, len(agg_headers) + 1):
        col_letter = get_column_letter(c_idx)
        total_formula = f"=SUM({col_letter}3:{col_letter}{data_end})"
        _data_cell(ws, data_end + 1, c_idx, total_formula, bold=True, align="right", fill=totals_fill)

    # Conditional formatting on Total Revenue column (col 4)
    if len(agg_headers) >= 4:
        rev_col = 4
        rule = ColorScaleRule(
            start_type="min", start_color=RED,
            mid_type="percentile", mid_value=50, mid_color=YELLOW,
            end_type="max", end_color=GREEN,
        )
        rev_col_letter = get_column_letter(rev_col)
        ws.conditional_formatting.add(f"{rev_col_letter}3:{rev_col_letter}{data_end}", rule)

    _autosize(ws)

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Revenue by Region"
    chart.style = 10
    chart.width = 24
    chart.height = 15
    chart.y_axis.title = "Revenue (USD)"
    chart.x_axis.title = "Region"

    if len(agg_headers) >= 4:
        data_ref = Reference(ws, min_col=4, min_row=2, max_row=data_end)
        cats_ref = Reference(ws, min_col=1, min_row=3, max_row=data_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        anchor_col = get_column_letter(len(agg_headers) + 2)
        ws.add_chart(chart, f"{anchor_col}2")


# ---------------------------------------------------------------------------
# Sheet 3: Detail
# ---------------------------------------------------------------------------

def build_detail_sheet(
    ws: Worksheet, headers: list[str], rows: list[list[Any]]
) -> None:
    """Full dataset with styled header, alternating rows, and color scale."""
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title = ws["A1"]
    title.value = "Transaction Detail"
    title.fill = PatternFill(fill_type="solid", fgColor=NAVY)
    title.font = Font(bold=True, color=WHITE, size=12)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Headers (row 2)
    for c_idx, h in enumerate(headers, start=1):
        _header_cell(ws, 2, c_idx, h.replace("_", " ").title(), bg=LIGHT_NAVY, size=10)
    ws.freeze_panes = ws["A3"]

    alt_fill = PatternFill(fill_type="solid", fgColor=ALT_ROW)
    for r_offset, row in enumerate(rows, start=1):
        r_idx = r_offset + 2
        row_fill = alt_fill if r_offset % 2 == 0 else None
        for c_idx, value in enumerate(row, start=1):
            num_fmt = "#,##0.00" if isinstance(value, float) else "General"
            _data_cell(
                ws, r_idx, c_idx, value,
                align="right" if isinstance(value, (int, float)) else "left",
                fill=row_fill,
                number_format=num_fmt,
            )

    # Color scale on revenue
    if "revenue" in headers:
        rev_col = headers.index("revenue") + 1
        rule = ColorScaleRule(
            start_type="min", start_color=RED,
            mid_type="percentile", mid_value=50, mid_color=YELLOW,
            end_type="max", end_color=GREEN,
        )
        rev_letter = get_column_letter(rev_col)
        ws.conditional_formatting.add(f"{rev_letter}3:{rev_letter}{len(rows) + 2}", rule)

    _autosize(ws)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a polished executive Excel workbook from a sales CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python executive_workbook.py --input sample_sales.csv --output exec_report.xlsx\n"
            "  python executive_workbook.py --input sample_sales.csv --output exec_report.xlsx "
            "--company 'Acme Corp'\n"
        ),
    )
    parser.add_argument("--input", "-i", required=True, type=Path, help="Input CSV file.")
    parser.add_argument("--output", "-o", required=True, type=Path, help="Output .xlsx file.")
    parser.add_argument(
        "--company", default="Your Company", help="Company name for the cover sheet."
    )
    parser.add_argument(
        "--report-title",
        default="Sales Performance Report",
        help="Report title for the cover sheet.",
    )
    parser.add_argument("--force", action="store_true", help="Overwrite output if it exists.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if args.output.exists() and not args.force:
        print(f"Error: {args.output} already exists. Use --force to overwrite.", file=sys.stderr)
        return 1

    headers, rows = load_csv(args.input)
    print(f"Loaded {len(rows)} rows × {len(headers)} columns from {args.input}")

    try:
        agg_headers, agg_rows = aggregate_by_region(headers, rows)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)

    cover_ws = wb.create_sheet("Cover")
    build_cover_sheet(cover_ws, args.company, args.report_title, headers, rows)
    print("Built 'Cover' sheet")

    summary_ws = wb.create_sheet("Summary")
    build_summary_sheet(summary_ws, agg_headers, agg_rows)
    print("Built 'Summary' sheet")

    detail_ws = wb.create_sheet("Detail")
    build_detail_sheet(detail_ws, headers, rows)
    print("Built 'Detail' sheet")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Executive workbook written → {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
