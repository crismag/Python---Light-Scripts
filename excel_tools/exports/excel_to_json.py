"""
excel_to_json.py — INTERMEDIATE
=================================
Convert a chosen sheet of an .xlsx workbook into a JSON file (records orientation).

Features:
  - Reads the first row as column headers
  - Skips entirely empty rows
  - Optional pretty-printing (--indent)
  - Optional field filtering (--fields col1 col2 ...)
  - Outputs to a file or stdout (--stdout)

Example commands:
    python excel_tools/exports/excel_to_json.py \\
        --input output/sales.xlsx \\
        --output output/sales.json \\
        --pretty

    python excel_tools/exports/excel_to_json.py \\
        --input output/sales.xlsx \\
        --sheet "Sheet1" \\
        --fields date region revenue \\
        --stdout
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Core extraction logic
# ---------------------------------------------------------------------------

def extract_sheet_records(
    workbook_path: Path,
    sheet_name: str | None,
    fields: list[str] | None,
) -> list[dict[str, Any]]:
    """Read a worksheet and return a list of record dicts.

    Args:
        workbook_path: Path to the .xlsx file.
        sheet_name:    Name of the sheet to read; if None, uses the active (first) sheet.
        fields:        If provided, only these columns are included in each record.

    Returns:
        List of dicts representing each non-empty data row.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    if sheet_name is None:
        ws = wb.active
    elif sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        raw_headers: tuple[Any, ...] = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]

    # Determine which column indices to keep
    if fields:
        missing = [f for f in fields if f not in headers]
        if missing:
            wb.close()
            raise ValueError(f"Requested fields not found in headers: {missing}. Headers: {headers}")
        keep_indices = [headers.index(f) for f in fields]
        out_headers = fields
    else:
        keep_indices = list(range(len(headers)))
        out_headers = headers

    records: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        # Skip entirely empty rows
        if all(v is None for v in raw_row):
            continue
        record: dict[str, Any] = {}
        for out_h, src_idx in zip(out_headers, keep_indices):
            val = raw_row[src_idx] if src_idx < len(raw_row) else None
            # Convert non-JSON-serialisable types (e.g. datetime) to strings
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            record[out_h] = val
        records.append(record)

    wb.close()
    return records


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export a sheet from an .xlsx workbook to JSON (records orientation).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python excel_to_json.py --input data.xlsx --output data.json --pretty\n"
            "  python excel_to_json.py --input data.xlsx --sheet Sales --fields date region --stdout\n"
        ),
    )
    parser.add_argument("--input", "-i", required=True, type=Path, help="Input .xlsx file path.")
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=None,
        help="Output JSON file path. Required unless --stdout is set.",
    )
    parser.add_argument(
        "--sheet", "-s",
        default=None,
        metavar="SHEET_NAME",
        help="Sheet name to export (default: first/active sheet).",
    )
    parser.add_argument(
        "--fields",
        nargs="+",
        default=None,
        metavar="COLUMN",
        help="Only include these columns in the output (space-separated).",
    )
    parser.add_argument(
        "--pretty", "-p",
        action="store_true",
        help="Pretty-print JSON with 2-space indentation.",
    )
    parser.add_argument(
        "--stdout",
        action="store_true",
        help="Print JSON to stdout instead of writing a file.",
    )
    parser.add_argument(
        "--force", action="store_true", help="Overwrite output file if it already exists."
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if not args.stdout and args.output is None:
        print("Error: either --output or --stdout is required.", file=sys.stderr)
        return 1

    if args.output and args.output.exists() and not args.force and not args.stdout:
        print(f"Error: {args.output} already exists. Use --force to overwrite.", file=sys.stderr)
        return 1

    try:
        records = extract_sheet_records(args.input, sheet_name=args.sheet, fields=args.fields)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    indent = 2 if args.pretty else None
    json_str = json.dumps(records, indent=indent, default=str)

    if args.stdout:
        print(json_str)
    else:
        args.output.parent.mkdir(parents=True, exist_ok=True)  # type: ignore[union-attr]
        args.output.write_text(json_str, encoding="utf-8")  # type: ignore[union-attr]
        print(
            f"Exported {len(records)} records → {args.output}"
            + (f"  (fields: {args.fields})" if args.fields else "")
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
