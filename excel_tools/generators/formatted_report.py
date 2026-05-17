"""
formatted_report.py — INTERMEDIATE
====================================
Build a formatted report sheet from a CSV dataset:
  - Styled, bold header row with colour fill
  - Auto-sized columns
  - Alternating row shading
  - Three-colour conditional format on a chosen numeric column
  - Totals row at the bottom (SUM for numeric columns)

Example commands:
    python excel_tools/generators/formatted_report.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/sales_report.xlsx \\
        --title "Q1-Q3 Sales Report"

    python excel_tools/generators/formatted_report.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/sales_report.xlsx \\
        --cf-column revenue \\
        --force
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Self-contained helpers
# ---------------------------------------------------------------------------

def _load_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Load CSV, coercing numeric strings to int/float."""
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        records = list(reader)
    if not records:
        raise ValueError(f"CSV file is empty: {path}")
    headers = list(records[0].keys())
    rows: list[list[Any]] = []
    for rec in records:
        row: list[Any] = []
        for h in headers:
            val: Any = rec[h]
            try:
                val = int(val)
            except ValueError:
                try:
                    val = float(val)
                except ValueError:
                    pass
            row.append(val)
        rows.append(row)
    return headers, rows


def _write_title(ws: Worksheet, title: str, num_cols: int) -> None:
    """Write a merged title cell spanning all columns in row 1."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=14, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _write_header(ws: Worksheet, headers: list[str], row: int) -> None:
    """Write styled column headers."""
    fill = PatternFill(fill_type="solid", fgColor="2F5496")
    font = Font(bold=True, color="FFFFFF")
    for c_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _write_data(ws: Worksheet, rows: list[list[Any]], start_row: int) -> None:
    """Write data rows with alternating shading."""
    alt_fill = PatternFill(fill_type="solid", fgColor="EDF2FB")
    for r_offset, row in enumerate(rows):
        r_idx = start_row + r_offset
        fill = alt_fill if r_offset % 2 == 1 else None
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if fill:
                cell.fill = fill


def _write_totals(ws: Worksheet, headers: list[str], data_start: int, data_end: int) -> int:
    """Append a totals row; returns the totals row number."""
    totals_row = data_end + 1
    thin = Side(style="thin")
    top_border = Border(top=thin)
    total_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    bold_font = Font(bold=True)

    for c_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(c_idx)
        # Check if the column has numeric data by sampling the data start cell
        sample = ws.cell(row=data_start, column=c_idx).value
        if isinstance(sample, (int, float)):
            formula = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
            cell = ws.cell(row=totals_row, column=c_idx, value=formula)
        else:
            cell = ws.cell(row=totals_row, column=c_idx, value="TOTAL" if c_idx == 1 else None)
        cell.border = top_border
        cell.fill = total_fill
        cell.font = bold_font
    return totals_row


def _autosize(ws: Worksheet, min_w: int = 8, max_w: int = 55) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def _apply_color_scale(ws: Worksheet, col_letter: str, start_row: int, end_row: int) -> None:
    """Red→Yellow→Green three-colour scale on a column range."""
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="63BE7B",
    )
    ws.conditional_formatting.add(f"{col_letter}{start_row}:{col_letter}{end_row}", rule)


# ---------------------------------------------------------------------------
# Core builder
# ---------------------------------------------------------------------------

def build_report(
    headers: list[str],
    rows: list[list[Any]],
    output_path: Path,
    title: str,
    cf_column: str | None,
) -> None:
    """Build and save the formatted report workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    HEADER_ROW = 2
    DATA_START = 3

    _write_title(ws, title, len(headers))
    _write_header(ws, headers, HEADER_ROW)
    _write_data(ws, rows, DATA_START)

    data_end = DATA_START + len(rows) - 1
    _write_totals(ws, headers, DATA_START, data_end)
    _autosize(ws)

    # Apply conditional formatting to the requested column (or last numeric column)
    cf_col_idx: int | None = None
    if cf_column:
        try:
            cf_col_idx = headers.index(cf_column) + 1
        except ValueError:
            print(
                f"Warning: column '{cf_column}' not found; skipping conditional formatting.",
                file=sys.stderr,
            )
    else:
        # Default: last numeric column
        for idx, _ in enumerate(headers, start=1):
            if isinstance(ws.cell(row=DATA_START, column=idx).value, (int, float)):
                cf_col_idx = idx

    if cf_col_idx is not None:
        _apply_color_scale(ws, get_column_letter(cf_col_idx), DATA_START, data_end)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a formatted Excel report from a CSV file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python formatted_report.py --input sample_sales.csv --output report.xlsx\n"
            "  python formatted_report.py --input data.csv --output report.xlsx "
            "--cf-column revenue --force\n"
        ),
    )
    parser.add_argument("--input", "-i", required=True, type=Path, help="Input CSV file.")
    parser.add_argument("--output", "-o", required=True, type=Path, help="Output .xlsx file.")
    parser.add_argument(
        "--title",
        default="Data Report",
        help="Report title displayed at the top of the sheet (default: 'Data Report').",
    )
    parser.add_argument(
        "--cf-column",
        default=None,
        metavar="COLUMN_NAME",
        help="Column name to apply colour-scale conditional formatting to. "
             "Defaults to the last numeric column.",
    )
    parser.add_argument(
        "--force", action="store_true", help="Overwrite output file if it already exists."
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if args.output.exists() and not args.force:
        print(f"Error: {args.output} already exists. Use --force to overwrite.", file=sys.stderr)
        return 1

    headers, rows = _load_csv(args.input)
    print(f"Loaded {len(rows)} rows × {len(headers)} columns from {args.input}")

    build_report(headers, rows, args.output, title=args.title, cf_column=args.cf_column)
    print(f"Formatted report written → {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
