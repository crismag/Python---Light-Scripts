"""
single_sheet_generator.py — BEGINNER
=====================================
Read a CSV or JSON file and write a styled single-sheet .xlsx workbook.

The script auto-detects the input format from the file extension.
JSON input must be a list of objects (array of records).

Example commands:
    python excel_tools/generators/single_sheet_generator.py \\
        --input excel_tools/examples/sample_sales.csv \\
        --output output/sales.xlsx

    python excel_tools/generators/single_sheet_generator.py \\
        --input excel_tools/examples/sample_employees.json \\
        --output output/employees.xlsx \\
        --sheet-name Employees \\
        --header-color 1F4E79
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Internal helpers (self-contained — not imported from helpers.py)
# ---------------------------------------------------------------------------

def _write_header_row(ws: Any, headers: list[str], bg_color: str) -> None:
    """Write a styled header row to the worksheet."""
    fill = PatternFill(fill_type="solid", fgColor=bg_color)
    font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws["A2"]  # freeze header row


def _autosize_columns(ws: Any, min_width: int = 8, max_width: int = 60) -> None:
    """Fit each column to its widest content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_csv(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Return (headers, rows) from a CSV file."""
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        headers = next(reader)
        rows = list(reader)
    return headers, rows


def load_json(path: Path) -> tuple[list[str], list[list[Any]]]:
    """Return (headers, rows) from a JSON array-of-objects file."""
    with path.open(encoding="utf-8") as fh:
        records: list[dict[str, Any]] = json.load(fh)
    if not isinstance(records, list) or not records:
        raise ValueError("JSON input must be a non-empty array of objects.")
    headers = list(records[0].keys())
    rows = [[rec.get(h) for h in headers] for rec in records]
    return headers, rows


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

def build_workbook(
    headers: list[str],
    rows: list[list[Any]],
    sheet_name: str,
    header_color: str,
) -> Workbook:
    """Create an .xlsx workbook with one styled sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet names: max 31 chars

    # Write header
    _write_header_row(ws, headers, bg_color=header_color)

    # Write data rows — coerce numeric strings from CSV to float/int where possible
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            # Try to coerce strings that look like numbers
            if isinstance(value, str):
                try:
                    value = int(value)  # noqa: PLW2901  (intentional re-bind)
                except ValueError:
                    try:
                        value = float(value)  # noqa: PLW2901
                    except ValueError:
                        pass
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Alternate row shading for readability
    light_fill = PatternFill(fill_type="solid", fgColor="EDF2FB")
    for r_idx in range(2, len(rows) + 2, 2):
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = light_fill

    _autosize_columns(ws)
    return wb


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert a CSV or JSON file into a styled single-sheet .xlsx workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python single_sheet_generator.py --input sample_sales.csv --output sales.xlsx\n"
            "  python single_sheet_generator.py --input sample_employees.json --output emp.xlsx\n"
        ),
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        type=Path,
        help="Input CSV or JSON file path.",
    )
    parser.add_argument(
        "--output", "-o",
        required=True,
        type=Path,
        help="Output .xlsx file path.",
    )
    parser.add_argument(
        "--sheet-name",
        default="Sheet1",
        help="Name for the worksheet (default: Sheet1).",
    )
    parser.add_argument(
        "--header-color",
        default="2F5496",
        help="Header row background colour as hex string, no '#' (default: 2F5496).",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite output file if it already exists.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    # Safety: refuse to overwrite unless --force
    if args.output.exists() and not args.force:
        print(
            f"Error: {args.output} already exists. Use --force to overwrite.",
            file=sys.stderr,
        )
        return 1

    # Guard: input must exist
    if not args.input.exists():
        print(f"Error: input file not found: {args.input}", file=sys.stderr)
        return 1

    # Detect format and load data
    suffix = args.input.suffix.lower()
    if suffix == ".csv":
        headers, rows = load_csv(args.input)
    elif suffix == ".json":
        headers, rows = load_json(args.input)
    else:
        print(f"Error: unsupported file extension '{suffix}'. Use .csv or .json.", file=sys.stderr)
        return 1

    print(f"Loaded {len(rows)} rows × {len(headers)} columns from {args.input}")

    # Build and save workbook
    wb = build_workbook(headers, rows, sheet_name=args.sheet_name, header_color=args.header_color)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Workbook written → {args.output}  ({len(rows)} data rows, sheet='{args.sheet_name}')")
    return 0


if __name__ == "__main__":
    sys.exit(main())
