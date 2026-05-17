"""
sheet_validator.py — INTERMEDIATE
===================================
Validate a sheet in an .xlsx workbook against a configurable rule set:

  1. Required columns are present
  2. Numeric columns contain only non-negative values
  3. Key columns have no empty (None) cells

Prints a structured validation report to stdout and exits with code 0 (pass)
or 1 (fail).  Optionally accepts a JSON rules file; if omitted, defaults are
inferred automatically from the worksheet.

Example commands:
    # Auto-infer rules (check all numeric columns non-negative, no empty cells):
    python excel_tools/validators/sheet_validator.py \\
        --input output/sales.xlsx

    # Explicit rules:
    python excel_tools/validators/sheet_validator.py \\
        --input output/sales.xlsx \\
        --required-cols date region product units unit_price revenue \\
        --nonneg-cols units unit_price revenue \\
        --nonempty-cols date region product
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class ValidationIssue:
    rule: str
    column: str
    row: int | None
    detail: str


@dataclass
class ValidationReport:
    sheet_name: str
    total_rows: int
    issues: list[ValidationIssue] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return len(self.issues) == 0

    def print_report(self) -> None:
        status = "PASS" if self.passed else "FAIL"
        border = "=" * 60
        print(border)
        print(f"Validation Report  [{status}]")
        print(f"  Sheet      : {self.sheet_name}")
        print(f"  Data rows  : {self.total_rows}")
        print(f"  Issues     : {len(self.issues)}")
        print(border)
        if self.issues:
            for i, issue in enumerate(self.issues, start=1):
                row_str = f" (row {issue.row})" if issue.row else ""
                print(f"  [{i:>3}] {issue.rule:<22} col='{issue.column}'{row_str} — {issue.detail}")
        else:
            print("  No issues found.")
        print(border)


# ---------------------------------------------------------------------------
# Validator
# ---------------------------------------------------------------------------

def _load_sheet(
    path: Path, sheet_name: str | None
) -> tuple[openpyxl.worksheet.worksheet.Worksheet, openpyxl.Workbook]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name is None:
        ws = wb.active
    elif sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not in workbook. Available: {wb.sheetnames}")
    return ws, wb


def _read_data(ws: Any) -> tuple[list[str], list[list[Any]]]:
    """Return (headers, data_rows) from the worksheet."""
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers: tuple[Any, ...] = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]
    data_rows = [list(r) for r in rows_iter if not all(v is None for v in r)]
    return headers, data_rows


def validate(
    path: Path,
    sheet_name: str | None,
    required_cols: list[str],
    nonneg_cols: list[str],
    nonempty_cols: list[str],
) -> ValidationReport:
    """Run validation rules against the worksheet and return a report."""
    ws, wb = _load_sheet(path, sheet_name)
    headers, data_rows = _read_data(ws)
    wb.close()

    report = ValidationReport(sheet_name=ws.title, total_rows=len(data_rows))

    # --- Rule 1: Required columns present ---
    for col_name in required_cols:
        if col_name not in headers:
            report.issues.append(ValidationIssue(
                rule="REQUIRED_COL_MISSING",
                column=col_name,
                row=None,
                detail=f"Column '{col_name}' not found in headers.",
            ))

    # Build a header → column-index map for subsequent rules
    col_idx: dict[str, int] = {h: i for i, h in enumerate(headers)}

    # --- Rule 2: Non-negative numeric columns ---
    for col_name in nonneg_cols:
        if col_name not in col_idx:
            # Already flagged by rule 1 if listed; skip silently here
            continue
        c = col_idx[col_name]
        for r_offset, row in enumerate(data_rows, start=2):  # row 1 = header
            val = row[c] if c < len(row) else None
            if val is None:
                continue
            if isinstance(val, (int, float)) and val < 0:
                report.issues.append(ValidationIssue(
                    rule="NEGATIVE_VALUE",
                    column=col_name,
                    row=r_offset,
                    detail=f"Value {val} is negative.",
                ))

    # --- Rule 3: No empty cells in key columns ---
    for col_name in nonempty_cols:
        if col_name not in col_idx:
            continue
        c = col_idx[col_name]
        for r_offset, row in enumerate(data_rows, start=2):
            val = row[c] if c < len(row) else None
            if val is None or (isinstance(val, str) and val.strip() == ""):
                report.issues.append(ValidationIssue(
                    rule="EMPTY_CELL",
                    column=col_name,
                    row=r_offset,
                    detail="Cell is empty or blank.",
                ))

    return report


# ---------------------------------------------------------------------------
# Auto-inference of rules when none are provided
# ---------------------------------------------------------------------------

def _infer_rules(
    path: Path, sheet_name: str | None
) -> tuple[list[str], list[str], list[str]]:
    """Return (required_cols, nonneg_cols, nonempty_cols) inferred from the sheet."""
    ws, wb = _load_sheet(path, sheet_name)
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers: tuple[Any, ...] = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [], []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]

    # Sample up to 10 rows to classify columns
    sample: list[tuple[Any, ...]] = []
    for i, row in enumerate(rows_iter):
        sample.append(row)
        if i >= 9:
            break
    wb.close()

    nonneg: list[str] = []
    nonempty: list[str] = []

    for c_idx, h in enumerate(headers):
        vals = [row[c_idx] for row in sample if c_idx < len(row) and row[c_idx] is not None]
        if vals and all(isinstance(v, (int, float)) for v in vals):
            nonneg.append(h)
        # First 3 columns are treated as key (non-empty) columns
        if c_idx < 3:
            nonempty.append(h)

    return headers, nonneg, nonempty


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Validate an .xlsx sheet: required columns, non-negative numerics, no empty key cells."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python sheet_validator.py --input data.xlsx\n"
            "  python sheet_validator.py --input data.xlsx "
            "--required-cols date region --nonneg-cols revenue\n"
        ),
    )
    parser.add_argument("--input", "-i", required=True, type=Path, help="Input .xlsx file.")
    parser.add_argument("--sheet", "-s", default=None, help="Sheet name (default: active sheet).")
    parser.add_argument(
        "--required-cols",
        nargs="*",
        default=None,
        metavar="COL",
        help="Column names that must be present. Inferred from sheet if omitted.",
    )
    parser.add_argument(
        "--nonneg-cols",
        nargs="*",
        default=None,
        metavar="COL",
        help="Numeric columns that must be >= 0. Inferred from sheet if omitted.",
    )
    parser.add_argument(
        "--nonempty-cols",
        nargs="*",
        default=None,
        metavar="COL",
        help="Columns that must not contain empty cells. Inferred from sheet if omitted.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if not args.input.exists():
        print(f"Error: input file not found: {args.input}", file=sys.stderr)
        return 1

    # Infer rules if not provided
    if args.required_cols is None or args.nonneg_cols is None or args.nonempty_cols is None:
        inferred_req, inferred_nonneg, inferred_nonempty = _infer_rules(args.input, args.sheet)
        required_cols = args.required_cols if args.required_cols is not None else inferred_req
        nonneg_cols = args.nonneg_cols if args.nonneg_cols is not None else inferred_nonneg
        nonempty_cols = args.nonempty_cols if args.nonempty_cols is not None else inferred_nonempty
        print("(Rules inferred from sheet — pass explicit flags to override)")
    else:
        required_cols = args.required_cols
        nonneg_cols = args.nonneg_cols
        nonempty_cols = args.nonempty_cols

    try:
        report = validate(
            args.input,
            sheet_name=args.sheet,
            required_cols=required_cols,
            nonneg_cols=nonneg_cols,
            nonempty_cols=nonempty_cols,
        )
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    report.print_report()
    return 0 if report.passed else 1


if __name__ == "__main__":
    sys.exit(main())
