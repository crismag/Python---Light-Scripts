"""
test_single_sheet_generator.py
================================
Integration tests for excel_tools/generators/single_sheet_generator.py.

Strategy:
  1. Write temporary CSV / JSON fixture files
  2. Run the generator's main() function against them
  3. Reopen the produced .xlsx with openpyxl and assert cell contents

Run with:
    python -m pytest excel_tools/tests/test_single_sheet_generator.py -v
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
import pytest

# Make generator importable by adding the repo root to sys.path
_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from excel_tools.generators.single_sheet_generator import main  # noqa: E402

# ---------------------------------------------------------------------------
# Fixtures: temporary CSV and JSON data
# ---------------------------------------------------------------------------

_CSV_CONTENT = """\
name,age,salary
Alice,30,75000
Bob,25,62000
Carol,35,88000
"""

_JSON_CONTENT = [
    {"product": "Widget A", "units": 120, "revenue": 1198.80},
    {"product": "Gadget Pro", "units": 40, "revenue": 1999.60},
    {"product": "Accessory X", "units": 150, "revenue": 637.50},
]


@pytest.fixture()
def csv_file(tmp_path: Path) -> Path:
    p = tmp_path / "data.csv"
    p.write_text(_CSV_CONTENT, encoding="utf-8")
    return p


@pytest.fixture()
def json_file(tmp_path: Path) -> Path:
    p = tmp_path / "data.json"
    p.write_text(json.dumps(_JSON_CONTENT), encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# CSV input tests
# ---------------------------------------------------------------------------

class TestCSVInput:
    def test_returns_zero_on_success(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        result = main(["--input", str(csv_file), "--output", str(out)])
        assert result == 0

    def test_output_file_created(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        assert out.exists()
        assert out.stat().st_size > 0

    def test_header_row_correct(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws["A1"].value == "name"
        assert ws["B1"].value == "age"
        assert ws["C1"].value == "salary"

    def test_data_rows_present(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws["A2"].value == "Alice"
        assert ws["A3"].value == "Bob"
        assert ws["A4"].value == "Carol"

    def test_numeric_coercion(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # age should be int, salary should be int
        assert isinstance(ws["B2"].value, (int, float))
        assert isinstance(ws["C2"].value, (int, float))

    def test_row_count(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # 1 header + 3 data rows
        assert ws.max_row == 4

    def test_custom_sheet_name(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out), "--sheet-name", "Employees"])
        wb = openpyxl.load_workbook(out)
        assert "Employees" in wb.sheetnames

    def test_refuses_overwrite_without_force(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        # Second call without --force should return 1
        result = main(["--input", str(csv_file), "--output", str(out)])
        assert result == 1

    def test_force_flag_allows_overwrite(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        result = main(["--input", str(csv_file), "--output", str(out), "--force"])
        assert result == 0

    def test_creates_parent_directories(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "nested" / "deep" / "output.xlsx"
        result = main(["--input", str(csv_file), "--output", str(out)])
        assert result == 0
        assert out.exists()

    def test_freeze_panes_set(self, csv_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(csv_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # freeze_panes should be set (row 2)
        assert ws.freeze_panes is not None


# ---------------------------------------------------------------------------
# JSON input tests
# ---------------------------------------------------------------------------

class TestJSONInput:
    def test_returns_zero_on_success(self, json_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        result = main(["--input", str(json_file), "--output", str(out)])
        assert result == 0

    def test_header_row_from_json_keys(self, json_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(json_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws["A1"].value == "product"
        assert ws["B1"].value == "units"
        assert ws["C1"].value == "revenue"

    def test_first_data_row(self, json_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(json_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws["A2"].value == "Widget A"
        assert ws["B2"].value == 120

    def test_row_count(self, json_file: Path, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        main(["--input", str(json_file), "--output", str(out)])
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # 1 header + 3 data rows
        assert ws.max_row == 4


# ---------------------------------------------------------------------------
# Error handling
# ---------------------------------------------------------------------------

class TestErrorHandling:
    def test_missing_input_returns_one(self, tmp_path: Path) -> None:
        out = tmp_path / "output.xlsx"
        result = main(["--input", str(tmp_path / "nonexistent.csv"), "--output", str(out)])
        assert result == 1

    def test_unsupported_extension_returns_one(self, tmp_path: Path) -> None:
        bad_file = tmp_path / "data.parquet"
        bad_file.write_bytes(b"fake")
        out = tmp_path / "output.xlsx"
        result = main(["--input", str(bad_file), "--output", str(out)])
        assert result == 1
