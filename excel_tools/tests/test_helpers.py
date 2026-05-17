"""
test_helpers.py
================
Unit tests for excel_tools/helpers.py.

Covers:
  - autosize_columns
  - write_header_row
  - apply_color_scale
  - write_records_to_sheet
  - add_bar_chart
  - build_demo_workbook (integration)

Run with:
    python -m pytest excel_tools/tests/test_helpers.py -v
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook

# Make helpers importable by adding the repo root to sys.path
_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from excel_tools.helpers import (  # noqa: E402
    add_bar_chart,
    apply_color_scale,
    autosize_columns,
    build_demo_workbook,
    write_header_row,
    write_records_to_sheet,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture()
def blank_wb() -> Workbook:
    return Workbook()


@pytest.fixture()
def blank_ws(blank_wb: Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    return blank_wb.active


# ---------------------------------------------------------------------------
# autosize_columns
# ---------------------------------------------------------------------------

class TestAutosizeColumns:
    def test_wider_content_increases_width(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        blank_ws["A1"] = "Short"
        blank_ws["A2"] = "A" * 40  # very long
        autosize_columns(blank_ws)
        assert blank_ws.column_dimensions["A"].width > 10

    def test_max_width_respected(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        blank_ws["A1"] = "X" * 200
        autosize_columns(blank_ws, max_width=30)
        assert blank_ws.column_dimensions["A"].width <= 30

    def test_min_width_respected(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        blank_ws["A1"] = "Hi"
        autosize_columns(blank_ws, min_width=15)
        assert blank_ws.column_dimensions["A"].width >= 15

    def test_empty_sheet_does_not_error(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        autosize_columns(blank_ws)  # should not raise


# ---------------------------------------------------------------------------
# write_header_row
# ---------------------------------------------------------------------------

class TestWriteHeaderRow:
    def test_values_written(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        headers = ["Name", "Age", "Score"]
        write_header_row(blank_ws, headers)
        assert blank_ws["A1"].value == "Name"
        assert blank_ws["B1"].value == "Age"
        assert blank_ws["C1"].value == "Score"

    def test_font_is_bold(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_header_row(blank_ws, ["Col"], bold=True)
        assert blank_ws["A1"].font.bold is True

    def test_custom_row(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_header_row(blank_ws, ["A", "B"], row=3)
        assert blank_ws["A3"].value == "A"
        assert blank_ws["A1"].value is None

    def test_freeze_panes_set(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_header_row(blank_ws, ["X"], row=1, freeze=True)
        # openpyxl stores freeze_panes as a cell coordinate string
        assert blank_ws.freeze_panes == "A2"

    def test_no_freeze(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_header_row(blank_ws, ["X"], freeze=False)
        assert blank_ws.freeze_panes is None


# ---------------------------------------------------------------------------
# apply_color_scale
# ---------------------------------------------------------------------------

class TestApplyColorScale:
    def test_conditional_format_added(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        for r in range(1, 6):
            blank_ws.cell(row=r, column=1, value=r * 10)
        apply_color_scale(blank_ws, col_letter="A", start_row=1, end_row=5)
        # openpyxl stores rules in conditional_formatting
        assert len(blank_ws.conditional_formatting._cf_rules) > 0

    def test_custom_colors_accepted(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        apply_color_scale(
            blank_ws, "B", 2, 10,
            low_color="FF0000", mid_color="00FF00", high_color="0000FF",
        )
        # No exception = pass


# ---------------------------------------------------------------------------
# write_records_to_sheet
# ---------------------------------------------------------------------------

class TestWriteRecordsToSheet:
    _RECORDS = [
        {"Name": "Alice", "Score": 95},
        {"Name": "Bob",   "Score": 87},
        {"Name": "Carol", "Score": 78},
    ]

    def test_header_written(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_records_to_sheet(blank_ws, self._RECORDS)
        assert blank_ws["A1"].value == "Name"
        assert blank_ws["B1"].value == "Score"

    def test_data_written(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_records_to_sheet(blank_ws, self._RECORDS)
        assert blank_ws["A2"].value == "Alice"
        assert blank_ws["B4"].value == 78

    def test_empty_records_no_error(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_records_to_sheet(blank_ws, [])  # should not raise

    def test_row_count(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_records_to_sheet(blank_ws, self._RECORDS)
        # 1 header + 3 data rows = max_row 4
        assert blank_ws.max_row == 4

    def test_custom_start_row(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        write_records_to_sheet(blank_ws, self._RECORDS, start_row=3)
        assert blank_ws["A3"].value == "Name"  # header at row 3
        assert blank_ws["A4"].value == "Alice"  # first data at row 4


# ---------------------------------------------------------------------------
# add_bar_chart
# ---------------------------------------------------------------------------

class TestAddBarChart:
    def test_chart_added_to_sheet(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        # Write minimal data
        data = [("Category", "Value"), ("A", 10), ("B", 20), ("C", 30)]
        for r_idx, (cat, val) in enumerate(data, start=1):
            blank_ws.cell(row=r_idx, column=1, value=cat)
            blank_ws.cell(row=r_idx, column=2, value=val)

        add_bar_chart(blank_ws, data_min_row=1, data_max_row=4, categories_col=1, values_col=2)
        assert len(blank_ws._charts) == 1

    def test_chart_title_set(self, blank_ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        for r in range(1, 4):
            blank_ws.cell(row=r, column=1, value=f"Cat{r}")
            blank_ws.cell(row=r, column=2, value=r * 5)
        add_bar_chart(
            blank_ws, data_min_row=1, data_max_row=3,
            categories_col=1, values_col=2, title="My Chart",
        )
        # openpyxl wraps chart titles in a Title object; extract the text string
        chart = blank_ws._charts[0]
        title_text = chart.title
        if hasattr(title_text, "tx"):
            # Title object — extract the first run text
            title_text = title_text.tx.rich.p[0].r[0].t
        assert title_text == "My Chart"


# ---------------------------------------------------------------------------
# build_demo_workbook — integration round-trip
# ---------------------------------------------------------------------------

class TestBuildDemoWorkbook:
    def test_file_created(self, tmp_path: Path) -> None:
        out = tmp_path / "demo.xlsx"
        build_demo_workbook(out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_workbook_valid(self, tmp_path: Path) -> None:
        out = tmp_path / "demo.xlsx"
        build_demo_workbook(out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws["A1"].value == "Region"

    def test_data_rows_present(self, tmp_path: Path) -> None:
        out = tmp_path / "demo.xlsx"
        build_demo_workbook(out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # 5 data rows + 1 header
        assert ws.max_row >= 6

    def test_chart_embedded(self, tmp_path: Path) -> None:
        out = tmp_path / "demo.xlsx"
        build_demo_workbook(out)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert len(ws._charts) >= 1
