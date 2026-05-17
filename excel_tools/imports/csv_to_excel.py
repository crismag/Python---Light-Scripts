"""
csv_to_excel.py — BEGINNER
============================
Convert one or more CSV files into a single .xlsx workbook.
Each CSV becomes one worksheet; the sheet name is derived from the CSV filename.

Example commands:
    # Single file:
    python excel_tools/imports/csv_to_excel.py \\
        --inputs excel_tools/examples/sample_sales.csv \\
        --output output/sales.xlsx

    # Multiple files (each becomes its own sheet):
    python excel_tools/imports/csv_to_excel.py \\
        --inputs data/q1.csv data/q2.csv data/q3.csv \\
        --output output/quarterly.xlsx \\
        --force
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Self-contained helpers
# ---------------------------------------------------------------------------

def _coerce(value: str) -> Any:
    """Attempt to coerce a string value to int or float, else return as-is."""
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        return value


def _safe_sheet_name(path: Path, existing: set[str]) -> str:
    """Derive a unique, Excel-safe sheet name from a file path."""
    stem = path.stem[:27]  # leave room for dedup suffix
    name = stem
    counter = 2
    while name in existing:
        name = f"{stem}_{counter}"
        counter += 1
    existing.add(name)
    return name


def _write_csv_to_sheet(ws: Any, csv_path: Path, header_color: str) -> int:
    """Read *csv_path* and write it into *ws*. Returns the number of data rows written."""
    with csv_path.open(newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        all_rows = list(reader)

    if not all_rows:
        return 0

    headers = all_rows[0]
    data_rows = all_rows[1:]

    # Styled header
    fill = PatternFill(fill_type="solid", fgColor=header_color)
    font = Font(bold=True, color="FFFFFF")
    for c_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws["A2"]

    # Data with numeric coercion
    alt_fill = PatternFill(fill_type="solid", fgColor="F2F7FF")
    for r_offset, row in enumerate(data_rows, start=1):
        r_idx = r_offset + 1
        row_fill = alt_fill if r_offset % 2 == 0 else None
        for c_idx, raw_val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_coerce(raw_val))
            if row_fill:
                cell.fill = row_fill

    # Auto-size columns
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=6)
        ws.column_dimensions[col_letter].width = max(8, min(max_len + 2, 55))

    return len(data_rows)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert one or more CSV files into a single .xlsx workbook (one sheet per CSV).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python csv_to_excel.py --inputs sales.csv --output workbook.xlsx\n"
            "  python csv_to_excel.py --inputs q1.csv q2.csv q3.csv --output quarterly.xlsx\n"
        ),
    )
    parser.add_argument(
        "--inputs", "-i",
        nargs="+",
        required=True,
        type=Path,
        metavar="CSV_FILE",
        help="One or more CSV files to import.",
    )
    parser.add_argument(
        "--output", "-o",
        required=True,
        type=Path,
        help="Output .xlsx file path.",
    )
    parser.add_argument(
        "--header-color",
        default="2F5496",
        metavar="HEX",
        help="Header row background colour as hex (no '#', default: 2F5496).",
    )
    parser.add_argument(
        "--force", action="store_true", help="Overwrite the output file if it already exists."
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if args.output.exists() and not args.force:
        print(f"Error: {args.output} already exists. Use --force to overwrite.", file=sys.stderr)
        return 1

    # Validate inputs
    missing = [p for p in args.inputs if not p.exists()]
    if missing:
        for p in missing:
            print(f"Error: input file not found: {p}", file=sys.stderr)
        return 1

    wb = Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    used_names: set[str] = set()
    for csv_path in args.inputs:
        sheet_name = _safe_sheet_name(csv_path, used_names)
        ws = wb.create_sheet(title=sheet_name)
        n_rows = _write_csv_to_sheet(ws, csv_path, header_color=args.header_color)
        print(f"  [{sheet_name}]  {n_rows} data rows  ← {csv_path}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Workbook written → {args.output}  ({len(args.inputs)} sheet(s))")
    return 0


if __name__ == "__main__":
    sys.exit(main())
