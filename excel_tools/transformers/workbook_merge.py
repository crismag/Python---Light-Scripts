"""
workbook_merge.py — INTERMEDIATE
==================================
Merge several .xlsx workbooks into a single output workbook.
Each sheet from each input workbook is copied into the output,
with sheet names disambiguated when conflicts arise.

Features:
  - Copies cell values, basic styles, column widths, and row heights
  - Preserves sheet names, adding a source-file prefix on collision
  - Never overwrites the output file unless --force is given
  - Optionally prefixes all sheet names with the source filename stem

Example commands:
    python excel_tools/transformers/workbook_merge.py \\
        --inputs output/sales.xlsx output/employees.xlsx \\
        --output output/merged.xlsx

    python excel_tools/transformers/workbook_merge.py \\
        --inputs output/sales.xlsx output/employees.xlsx \\
        --output output/merged.xlsx \\
        --prefix-sheets \\
        --force
"""

from __future__ import annotations

import argparse
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook

# ---------------------------------------------------------------------------
# Sheet copying
# ---------------------------------------------------------------------------

def _unique_name(candidate: str, used: set[str], max_len: int = 31) -> str:
    """Return a sheet name that is not already in *used*, truncated to *max_len*."""
    base = candidate[:max_len]
    name = base
    counter = 2
    while name in used:
        suffix = f"_{counter}"
        name = base[: max_len - len(suffix)] + suffix
        counter += 1
    used.add(name)
    return name


def _copy_sheet(src_ws: openpyxl.worksheet.worksheet.Worksheet, dst_wb: Workbook, name: str) -> None:
    """Copy cell values and basic formatting from *src_ws* into a new sheet in *dst_wb*."""
    dst_ws = dst_wb.create_sheet(title=name)

    # Copy column dimensions
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row dimensions
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height
        dst_ws.row_dimensions[row_num].hidden = dim.hidden

    # Copy cells: value + style
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column, value=src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format

    # Copy merged cell ranges
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # Freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Merge several .xlsx workbooks into one. "
            "Each source sheet is copied into the output workbook."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python workbook_merge.py --inputs a.xlsx b.xlsx --output merged.xlsx\n"
            "  python workbook_merge.py --inputs *.xlsx --output merged.xlsx --prefix-sheets\n"
        ),
    )
    parser.add_argument(
        "--inputs", "-i",
        nargs="+",
        required=True,
        type=Path,
        metavar="XLSX_FILE",
        help="Two or more .xlsx files to merge.",
    )
    parser.add_argument(
        "--output", "-o",
        required=True,
        type=Path,
        help="Output .xlsx file path.",
    )
    parser.add_argument(
        "--prefix-sheets",
        action="store_true",
        help="Prefix every sheet name with the source file stem (e.g. 'sales_Sheet1').",
    )
    parser.add_argument(
        "--force", action="store_true", help="Overwrite output file if it already exists."
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if args.output.exists() and not args.force:
        print(f"Error: {args.output} already exists. Use --force to overwrite.", file=sys.stderr)
        return 1

    missing = [p for p in args.inputs if not p.exists()]
    if missing:
        for p in missing:
            print(f"Error: input file not found: {p}", file=sys.stderr)
        return 1

    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)  # remove default blank sheet

    used_names: set[str] = set()
    total_sheets = 0

    for src_path in args.inputs:
        src_wb = openpyxl.load_workbook(src_path, data_only=True)
        stem = src_path.stem

        for sheet_name in src_wb.sheetnames:
            candidate = f"{stem}_{sheet_name}" if args.prefix_sheets else sheet_name
            final_name = _unique_name(candidate, used_names)
            src_ws = src_wb[sheet_name]
            _copy_sheet(src_ws, dst_wb, final_name)
            print(f"  Copied [{sheet_name}] from {src_path.name}  →  [{final_name}]")
            total_sheets += 1

        src_wb.close()

    args.output.parent.mkdir(parents=True, exist_ok=True)
    dst_wb.save(args.output)
    print(f"Merged workbook written → {args.output}  ({total_sheets} sheets from {len(args.inputs)} files)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
