"""
helpers.py — standalone reference module for reusable openpyxl utility functions.

This file is a STUDY / COPY REFERENCE — the example scripts in this section do NOT
import it.  Copy individual functions into your own script as needed.

Usage (standalone demo):
    python excel_tools/helpers.py                    # writes helpers_demo.xlsx
    python excel_tools/helpers.py --out /tmp/demo.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# 1. Auto-size columns to content
# ---------------------------------------------------------------------------

def autosize_columns(ws: Worksheet, min_width: int = 8, max_width: int = 60) -> None:
    """Adjust each column's width to fit its widest cell value.

    Args:
        ws:         The openpyxl worksheet to resize.
        min_width:  Minimum column width (characters). Default 8.
        max_width:  Maximum column width (characters). Default 60.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ---------------------------------------------------------------------------
# 2. Write a styled header row
# ---------------------------------------------------------------------------

def write_header_row(
    ws: Worksheet,
    headers: list[str],
    row: int = 1,
    font_color: str = "FFFFFF",
    bg_color: str = "2F5496",
    bold: bool = True,
    freeze: bool = True,
) -> None:
    """Write column headers with a coloured background and optional freeze.

    Args:
        ws:          Target worksheet.
        headers:     List of header strings (written left-to-right from column A).
        row:         Row number to write into (1-indexed). Default 1.
        font_color:  Hex colour for header text (no leading '#'). Default white.
        bg_color:    Hex fill colour for header cells. Default navy blue.
        bold:        Whether to bold the header text. Default True.
        freeze:      If True, freeze the row below the header. Default True.
    """
    fill = PatternFill(fill_type="solid", fgColor=bg_color)
    font = Font(bold=bold, color=font_color)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if freeze:
        # Freeze the row immediately below the header row
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


# ---------------------------------------------------------------------------
# 3. Apply a conditional-format colour scale to a range
# ---------------------------------------------------------------------------

def apply_color_scale(
    ws: Worksheet,
    col_letter: str,
    start_row: int,
    end_row: int,
    low_color: str = "F8696B",   # red
    mid_color: str = "FFEB84",   # yellow
    high_color: str = "63BE7B",  # green
) -> None:
    """Apply a three-colour scale conditional format to a column range.

    Args:
        ws:          Target worksheet.
        col_letter:  Column letter (e.g. 'E').
        start_row:   First data row (1-indexed).
        end_row:     Last data row (inclusive).
        low_color:   Hex colour for the minimum value. Default red.
        mid_color:   Hex colour for the midpoint. Default yellow.
        high_color:  Hex colour for the maximum value. Default green.
    """
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    rule = ColorScaleRule(
        start_type="min",
        start_color=low_color,
        mid_type="percentile",
        mid_value=50,
        mid_color=mid_color,
        end_type="max",
        end_color=high_color,
    )
    ws.conditional_formatting.add(cell_range, rule)


# ---------------------------------------------------------------------------
# 4. Write a list-of-dicts as a sheet
# ---------------------------------------------------------------------------

def write_records_to_sheet(
    ws: Worksheet,
    records: list[dict[str, Any]],
    header_bg: str = "2F5496",
    start_row: int = 1,
) -> None:
    """Write a list of dicts to *ws* starting at *start_row*.

    The keys of the first record are used as column headers.

    Args:
        ws:         Target worksheet (will be written in-place).
        records:    List of dicts; all dicts should share the same keys.
        header_bg:  Background colour for the header row. Default navy.
        start_row:  Row to begin writing (1-indexed). Default 1.
    """
    if not records:
        return
    headers = list(records[0].keys())
    write_header_row(ws, headers, row=start_row, bg_color=header_bg)
    for r_offset, record in enumerate(records, start=1):
        for c_idx, key in enumerate(headers, start=1):
            ws.cell(row=start_row + r_offset, column=c_idx, value=record.get(key))
    autosize_columns(ws)


# ---------------------------------------------------------------------------
# 5. Add a simple bar chart to a sheet
# ---------------------------------------------------------------------------

def add_bar_chart(
    ws: Worksheet,
    data_min_row: int,
    data_max_row: int,
    categories_col: int,
    values_col: int,
    title: str = "Chart",
    anchor_cell: str = "H2",
    width: int = 20,
    height: int = 14,
) -> None:
    """Add a vertical bar chart to *ws* reading from an existing data block.

    Args:
        ws:              Worksheet that contains the data AND will host the chart.
        data_min_row:    First data row (1-indexed, including header if you want a legend label).
        data_max_row:    Last data row (inclusive).
        categories_col:  Column index (1-indexed) for X-axis category labels.
        values_col:      Column index (1-indexed) for the numeric bar series.
        title:           Chart title string.
        anchor_cell:     Top-left cell where the chart is placed (e.g. 'H2').
        width:           Chart width in cm. Default 20.
        height:          Chart height in cm. Default 14.
    """
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    data_ref = Reference(ws, min_col=values_col, min_row=data_min_row, max_row=data_max_row)
    cats_ref = Reference(ws, min_col=categories_col, min_row=data_min_row + 1, max_row=data_max_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, anchor_cell)


# ---------------------------------------------------------------------------
# Standalone demo — builds a small workbook exercising every helper above
# ---------------------------------------------------------------------------

_DEMO_RECORDS = [
    {"Region": "North",  "Q1 Sales": 48200, "Q2 Sales": 51000, "Q3 Sales": 39800},
    {"Region": "South",  "Q1 Sales": 36700, "Q2 Sales": 42100, "Q3 Sales": 44500},
    {"Region": "East",   "Q1 Sales": 55900, "Q2 Sales": 49300, "Q3 Sales": 61200},
    {"Region": "West",   "Q1 Sales": 29400, "Q2 Sales": 33800, "Q3 Sales": 37600},
    {"Region": "Central","Q1 Sales": 41100, "Q2 Sales": 38900, "Q3 Sales": 43000},
]


def build_demo_workbook(out_path: Path) -> None:
    """Build a small demonstration workbook and save to *out_path*."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Demo"

    # Write records with styled header
    write_records_to_sheet(ws, _DEMO_RECORDS, header_bg="1F4E79")

    # Apply colour scale to Q1 Sales column (column B, rows 2-6)
    apply_color_scale(ws, col_letter="B", start_row=2, end_row=len(_DEMO_RECORDS) + 1)

    # Add a bar chart anchored at column F
    add_bar_chart(
        ws,
        data_min_row=1,
        data_max_row=len(_DEMO_RECORDS) + 1,
        categories_col=1,   # Region
        values_col=2,        # Q1 Sales
        title="Q1 Sales by Region",
        anchor_cell="F2",
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Demo workbook written → {out_path}")

    # Verify it round-trips correctly
    wb2 = openpyxl.load_workbook(out_path)
    ws2 = wb2.active
    if ws2["A1"].value != "Region":
        raise RuntimeError("Header sanity check failed: expected 'Region' in A1")
    print("Round-trip verification passed.")


def _parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="helpers.py demo — build a small .xlsx exercising every helper function.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path("helpers_demo.xlsx"),
        help="Output file path (default: helpers_demo.xlsx)",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite the output file if it already exists.",
    )
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = _parse_args()
    if args.out.exists() and not args.force:
        print(f"Error: {args.out} already exists. Use --force to overwrite.", file=sys.stderr)
        sys.exit(1)
    build_demo_workbook(args.out)
